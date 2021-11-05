from pprint import pprint

import openpyxl
import json
import re

excel_file = openpyxl.load_workbook('Ведомость остатков ОС,НМА, НПА.xlsx')
# {
#   "Numenklatures":[
namelist = excel_file.sheetnames  # ['TDSheet'] - получение название листа
employees_sheet = excel_file[str(namelist[0])]
currently_active_sheet = excel_file.active

maxRow = employees_sheet.max_row
maxColumn = employees_sheet.max_column
listName = list()
strCode = list()
twoStolbesh = list()
f = 0
collposit = list()
collp = 0
shagi1 = 0
shagi2 = 0
shagi = 10
shetchik = 10

for i in range(maxRow - 9):
    i = i + 10
    i = str(i)
    m = employees_sheet["C" + i].value
    twoStolbesh.append(str(m))
    n = employees_sheet["A" + i].value
    strCode.append(str(n))

for k in range(len(twoStolbesh)):
    listName.append(("list" + str(k)))  # создаем массив всех элементов
    listName[k] = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]

for j in range(len(twoStolbesh)):
    if str(twoStolbesh[j]) != 'None':
        listName[f][0] = str(f + 1)
        listName[f][4] = twoStolbesh[j]
        listName[f][5] = employees_sheet["J" + str(j + 10)].value
        listName[f][6] = employees_sheet["K" + str(j + 10)].value
        listName[f][7] = employees_sheet["M" + str(j + 10)].value
        listName[f][8] = employees_sheet["O" + str(j + 10)].value
        listName[f][9] = employees_sheet["P" + str(j + 10)].value
        listName[f][10] = employees_sheet["Q" + str(j + 10)].value
        listName[f][11] = str(employees_sheet["R" + str(j + 10)].value)
        listName[f][12] = str(employees_sheet["S" + str(j + 10)].value)
        listName[f][13] = str(employees_sheet["T" + str(j + 10)].value)
        listName[f][14] = str(employees_sheet["U" + str(j + 10)].value)
        listName[f][15] = str(employees_sheet["V" + str(j + 10)].value)
        listName[f][16] = str(employees_sheet["W" + str(j + 10)].value)
        listName[f][17] = str(employees_sheet["Y" + str(j + 10)].value)
        f = f + 1

for i, e in reversed(list(enumerate(listName))):
    if str(listName[i][5]) == '5':
        del listName[i]

for m in range(len(strCode)):
    k = str(strCode[m])
    if not re.search(r'\d{2}.\d{2}\D{3}', strCode[m]) is None:
        collposit.append(collp)
        collposit.append((strCode[m]))
        collp = 0
    else:
        collp = collp + 1

# Добавляем в номенклатуры 10.34(эти номера)
for i in range(2, len(collposit), 2):
    shetchik = shetchik + int(collposit[i])
    strdelprobel = str(employees_sheet["A" + str(shetchik)].value)
    delprobel = strdelprobel.split()
    delprobel = ''.join(delprobel)
    shagi1 = int(delprobel) + int(shagi1)
    shagi1 = shagi1 - shagi2
    shetchik = int(shetchik) + 1
    for j in range(shagi2, shagi1):
        listName[j][1] = collposit[i - 1]
    shagi2 = shagi1

# Добавляем в последние номенклатуры 10.34(эти номера)
lastposishion = len(listName)
for k in range(shagi2, lastposishion):
    listName[k][1] = collposit[-1]

twoStolbesh1 = list()
twoStolbesh2 = list()
# С отделом и ФИО: Берем ячейку с номенклатурой и берем ячейку слева выше и проверяем ее на число, если это не число то записываем к этой номеклатуре эту клетку и над ней клетку, если число то идем выше, пока не найдем не число, если над не числом находится число, то номенклатуре добавляем прочерк.
for i in range(maxRow - 9):
    i = i + 10
    i = str(i)
    m = employees_sheet["C" + i].value
    if str(m) != 'None':
        naum = employees_sheet["A" + i].value
        if not re.search(r'\d', str(naum)) is None:
            l = int(i) - 1
            otdel = employees_sheet["A" + str(l)].value
            if re.search(r'^\d', str(otdel)) is None:
                otdelZap = otdel
                twoStolbesh1.append(otdelZap)
            else:
                twoStolbesh1.append(otdelZap)
            l = l - 1
            NameFIO = employees_sheet["A" + str(l)].value
            if re.search(r'^\d', str(otdel)) is None:
                NameFIOZap = NameFIO
                twoStolbesh2.append(NameFIOZap)
            else:
                twoStolbesh2.append(NameFIOZap)
for i in range(len(twoStolbesh2)):
    if not re.search(r'\d', str(twoStolbesh2[i])) is None:
        twoStolbesh2[i] = '-'

for m in range(len(listName)):
    listName[m][3] = twoStolbesh1[m]
    listName[m][2] = twoStolbesh2[m]

# Запись в json
for k in range(len(listName)):
    NameStolb = [
        "ID",
        "Категория",
        "ФИО ответственного",
        "Отдел",
        "Основное средство",
        "Инвентарный номер",
        "ОКОФ",
        "Амортизационная группа",
        "Способ начисления амортизации",
        "Дата ввода в эксплуатацию",
        "Состояние",
        "Срок полезного использования",
        "Мес. норма износа",
        "Износ",
        "Балансовая стоимость",
        "Количество",
        "Сумма амортизации",
        "Остаточная стоимость",
    ]
    fruit_dictionary = dict(zip(NameStolb, listName[k]))
    with open("test.json", "a") as write_file:
        json.dump(fruit_dictionary, write_file, ensure_ascii=False, indent=2) # Запись в json