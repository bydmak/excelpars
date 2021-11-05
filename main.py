import os
from pprint import pprint
import openpyxl
import json
import re
from PyQt5.QtCore import QSize, Qt
from PyQt5.QtWidgets import QApplication, QMainWindow, QGridLayout, QWidget, QTableWidget, QTableWidgetItem, QHeaderView
from PyQt5 import QtWidgets
from PyQt5 import QtGui
from PyQt5 import QtCore
from mydesign import Ui_MainWindow  # импорт нашего сгенерированного файла
import sys

class mywindow(QtWidgets.QMainWindow):
    def __init__(self):# Обязательно нужно вызвать метод супер класса
        QMainWindow.__init__(self)

        self.setMinimumSize(QSize(800, 500))             # Устанавливаем размеры
        self.setWindowTitle("2C")    # Устанавливаем заголовок окна
        central_widget = QWidget(self)                  # Создаём центральный виджет
        self.setCentralWidget(central_widget)  # Устанавливаем центральный виджет
        self.pushButton = QtWidgets.QPushButton(self)
        self.pushButton.setGeometry(QtCore.QRect(30, 30, 201, 51))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("background-color: yellow;\n"
                                      "border-radius: 20%")
        self.pushButton.setObjectName("pushButton")
        self.pushButton.setText("Выбрать таблицу")
        self.pushButton.clicked.connect(self.buttonClicked)

        grid_layout = QGridLayout()             # Создаём QGridLayout
        central_widget.setLayout(grid_layout)   # Устанавливаем данное размещение в центральный виджет
        excel_file = openpyxl.load_workbook('Tb1.xlsx')

        namelist = excel_file.sheetnames  # ['TDSheet'] - получение название листа
        employees_sheet = excel_file[str(namelist[0])]
        currently_active_sheet = excel_file.active

        maxRow = employees_sheet.max_row
        maxColumn = employees_sheet.max_column
        listName = list()
        strCode = list()
        twoStolbesh = list()
        f = 0
        collposit = list()
        collp = 0
        shagi1 = 0
        shagi2 = 0
        shagi = 10
        shetchik = 10

        for i in range(maxRow - 9):
            i = i + 10
            i = str(i)
            m = employees_sheet["C" + i].value
            twoStolbesh.append(str(m))
            n = employees_sheet["A" + i].value
            strCode.append(str(n))

        for k in range(len(twoStolbesh)):
            listName.append(("list" + str(k)))  # создаем массив всех элементов
            listName[k] = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]

        for j in range(len(twoStolbesh)):
            if str(twoStolbesh[j]) != 'None':
                listName[f][0] = str(f + 1)
                listName[f][5] = twoStolbesh[j]
                listName[f][6] = employees_sheet["J" + str(j + 10)].value
                listName[f][7] = employees_sheet["K" + str(j + 10)].value
                listName[f][8] = employees_sheet["M" + str(j + 10)].value
                listName[f][9] = employees_sheet["O" + str(j + 10)].value
                listName[f][10] = employees_sheet["P" + str(j + 10)].value
                listName[f][11] = employees_sheet["Q" + str(j + 10)].value
                listName[f][12] = str(employees_sheet["R" + str(j + 10)].value)
                listName[f][13] = str(employees_sheet["S" + str(j + 10)].value)
                listName[f][14] = str(employees_sheet["T" + str(j + 10)].value)
                listName[f][15] = str(employees_sheet["U" + str(j + 10)].value)
                listName[f][16] = str(employees_sheet["V" + str(j + 10)].value)
                listName[f][17] = str(employees_sheet["W" + str(j + 10)].value)
                listName[f][18] = str(employees_sheet["Y" + str(j + 10)].value)
                f = f + 1

        for i, e in reversed(list(enumerate(listName))):
            if str(listName[i][5]) == '5':
                del listName[i]

        for m in range(len(strCode)):
            k = str(strCode[m])
            if not re.search(r'\d{2}.\d{2}\D{3}', strCode[m]) is None:
                collposit.append(collp)
                collposit.append((strCode[m]))
                collp = 0
            else:
                collp = collp + 1

        # Добавляем в номенклатуры 10.34(эти номера)
        for i in range(2, len(collposit), 2):
            shetchik = shetchik + int(collposit[i])
            strdelprobel = str(employees_sheet["A" + str(shetchik)].value)
            delprobel = strdelprobel.split()
            delprobel = ''.join(delprobel)
            shagi1 = int(delprobel) + int(shagi1)
            shagi1 = shagi1 - shagi2
            shetchik = int(shetchik) + 1
            for j in range(shagi2, shagi1):
                listName[j][1] = collposit[i - 1]
            shagi2 = shagi1

        # Добавляем в последние номенклатуры 10.34(эти номера)
        lastposishion = len(listName)
        for k in range(shagi2, lastposishion):
            listName[k][1] = collposit[-1]

        # От сюда нужно получить сначала 101.32 или типо того, затем получить количество номенклатур входящих в неё, далее такому количеству по порядку изменить первое поле на это название.
        # С отделом и ФИО: Берем ячейку с номенклатурой и берем ячейку слева выше и проверяем ее на число, если это не число то записываем к этой номеклатуре эту клетку и над ней клетку, если число то идем выше, пока не найдем не число, если над не числом находится число, то номенклатуре добавляем прочерк.

        # Запись в json
        for k in range(len(listName)):
            NameStolb = [
                "ID",
                "Категория",
                "Номер",
                "ФИО ответственного",
                "Отдел",
                "Основное средство",
                "Инвентарный номер",
                "ОКОФ",
                "Амортизационная группа",
                "Способ начисления амортизации",
                "Дата ввода в эксплуатацию",
                "Состояние",
                "Срок полезного использования",
                "Мес. норма износа",
                "Износ",
                "Балансовая стоимость",
                "Количество",
                "Сумма амортизации",
                "Остаточная стоимость",
            ]
            fruit_dictionary = dict(zip(NameStolb, listName[k]))
            with open("data_file.json", "a") as write_file:
                json.dump(fruit_dictionary, write_file, ensure_ascii=False, indent=2) # Запись в json


        table = QTableWidget(self)  # Создаём таблицу
        table.setColumnCount(18)     # Устанавливаем  колонки
        table.setRowCount(len(listName))        # строки в таблицу
        # Устанавливаем заголовки таблицы
        table.setHorizontalHeaderLabels([NameStolb[1], NameStolb[2], NameStolb[3], NameStolb[4], NameStolb[5], NameStolb[6], NameStolb[7], NameStolb[8], NameStolb[9], NameStolb[10], NameStolb[11], NameStolb[12], NameStolb[13], NameStolb[14], NameStolb[15], NameStolb[16], NameStolb[17], NameStolb[18]])
        # # Устанавливаем всплывающие подсказки на заголовки
        # table.horizontalHeaderItem(0).setToolTip("Column 1 ")
        # table.horizontalHeaderItem(1).setToolTip("Column 2 ")
        # table.horizontalHeaderItem(2).setToolTip("Column 3 ")

        # Устанавливаем выравнивание на заголовки
        table.horizontalHeaderItem(0).setTextAlignment(Qt.AlignHCenter)
        table.horizontalHeaderItem(1).setTextAlignment(Qt.AlignHCenter)
        table.horizontalHeaderItem(2).setTextAlignment(Qt.AlignHCenter)
        # заполняем первую строку
        for i in range(len(listName)):
            for j in range(18):
                table.setItem(i, j, QTableWidgetItem(listName[i][j+1]))


        # делаем ресайз колонок по содержимому
        table.resizeColumnsToContents()

        grid_layout.addWidget(table, 0, 0)   # Добавляем таблицу в сетку
        grid_layout.setContentsMargins(30, 100, 30, 30)


    def buttonClicked(self):
        file, _ = QtWidgets.QFileDialog.getOpenFileName(self,
                                                        'Open File',
                                                        './',
                                                        'py Files (*.xlsx)')
        print(type(file))
        excel_file = openpyxl.load_workbook(file)
        namelist = excel_file.sheetnames  # ['TDSheet'] - получение название листа
        employees_sheet = excel_file[str(namelist[0])]
        currently_active_sheet = excel_file.active
        print(employees_sheet["A" + '50'].value)
        if not file:
            return

app = QtWidgets.QApplication([])
app.setWindowIcon(QtGui.QIcon('icon1.png'))

application = mywindow()
application.setWindowIcon(QtGui.QIcon('icon1.png'))
application.show()
sys.exit(app.exec())
