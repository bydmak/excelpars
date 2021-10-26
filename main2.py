import openpyxl
import re

excel_file = openpyxl.load_workbook('Table2.xlsx')

# print(excel_file.sheetnames) # ['TDSheet'] - получение название листа
employees_sheet = excel_file['TDSheet']
currently_active_sheet = excel_file.active

maxRow = employees_sheet.max_row
maxColumn = employees_sheet.max_column
listName = list()
strCode = list()
twoStolbesh = list()

for i in range(maxRow - 9):
    i = i + 9
    i = str(i)
    m = employees_sheet["A" + i].value
    strCode.append(str(m))  # - заполняем первый столбец

for i in range(maxRow - 9):
    i = i + 9
    i = str(i)
    m = employees_sheet["C" + i].value
    twoStolbesh.append(str(m))
    if str(i) == str(maxRow-1):
        for j in range(6):
            twoStolbesh.append("del")

for k in range(len(strCode)):
    listName.append(("list" + str(k)))  # создаем массив всех элементов
    listName[k] = [1, 2, 3, 4, 5]
f = 0
for j in range(len(strCode)):
    if not re.search(r'\d{2}.\d{2}\D{3}', strCode[j]) is None:  # делаем нужную выборку
        listName[f][0] = strCode[j]
        listName[f][1] = employees_sheet["A" + str(j + 11)].value
        listName[f][2] = employees_sheet["A" + str(j + 12)].value
        listName[f][3] = employees_sheet["A" + str(j + 13)].value
        print(strCode[j])
        print(twoStolbesh[j+4])

        f = f + 1

for i, e in reversed(list(enumerate(listName))):
    if str(listName[i][0]) == '1':
        del listName[i]
print(twoStolbesh)
print(listName)